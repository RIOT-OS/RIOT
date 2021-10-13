#!/usr/bin/env python3

# Copyright (C) 2020 Inria
#
# This file is subject to the terms and conditions of the GNU Lesser
# General Public License v2.1. See the file LICENSE in the top level
# directory for more details.

import os
import argparse
import datetime
import warnings

from openpyxl import load_workbook
from jinja2 import FileSystemLoader, Environment


CURRENT_DIR = os.path.dirname(os.path.abspath(__file__))
RIOTBASE = os.getenv(
    "RIOTBASE", os.path.abspath(os.path.join(CURRENT_DIR, "../../../..")))
STM32_KCONFIG_DIR = os.path.join(RIOTBASE, "cpu/stm32/kconfigs")
STM32_VENDOR_DIR = os.path.join(RIOTBASE, "cpu/stm32/include/vendor/cmsis")
MODEL_COLUMN = 1


def parse_sheet(cpu_fam, sheets):
    """Parse the Excel sheet and return a dict."""
    models = []
    for sheet in sheets:
        # filter warning raised by openpyxl
        with warnings.catch_warnings(record=True):
            warnings.simplefilter("always")
            # Load the content of the xlsx sheet
            sheet = load_workbook(filename=sheet, data_only=True).active

        # Extract models from sheet
        for idx, row in enumerate(sheet.rows):
            # Row index starts at 1
            model = str(sheet.cell(row=idx + 1, column=MODEL_COLUMN).value)
            if not model.startswith("STM32"):
                continue
            models.append(model.replace("-", "_"))
    return sorted(models)


def parse_cpu_lines(cpu_fam):
    """Return the list of available CPU lines."""
    headers_dir = os.path.join(STM32_VENDOR_DIR, cpu_fam, "Include")
    cpu_lines = [
        header[:-2].upper() for header in os.listdir(headers_dir)
        if (
            header.startswith("stm32") and
            header != "stm32{}xx.h".format(cpu_fam)
        )
    ]
    return sorted(cpu_lines)


def _match(model, line):
    """Return True if a cpu model matches a cpu line, False otherwise."""
    model = model.replace("_", "")
    family_model = model[6:9]
    family_model_letter1 = model[9]
    family_model_letter2 = model[10] if len(model) >= 11 else None
    family_model_letter3 = model[11] if len(model) == 12 else None

    family_line = line[6:9]
    family_line_letter1 = line[9]
    family_line_letter2 = line[10] if len(line) >= 11 else None
    family_line_letter3 = line[11] if len(line) == 12 else None

    if family_line_letter1 == "X" and family_line_letter2 == "X":
        letters_match = True
    elif family_line_letter1 == "X":
        if family_model_letter3 is not None or family_line_letter3 is not None:
            letters_match = (
                (family_line_letter2 == family_model_letter2) and
                (family_line_letter3 == family_model_letter3)
            )
        else:
            letters_match = family_line_letter2 == family_model_letter2
    elif family_line_letter2 == "X":
        letters_match = family_line_letter1 == family_model_letter1
    else:
        letters_match = False
    return family_model == family_line and letters_match


def get_context(cpu_fam, models, lines):
    """Return a dict where keys are the models and values are the lines/fam."""
    mapping = []
    for model in models:
        found_line = False
        for line in lines:
            if _match(model, line):
                mapping.append(
                    {"model": model, "line": "CPU_LINE_{}".format(line)}
                )
                found_line = True
        # if a model has no matching line, just match it to the upper cpu
        # fam level
        if not found_line:
            mapping.append(
                {"model": model, "line": "CPU_FAM_{}".format(cpu_fam)}
            )
    return {"models": mapping, "lines": lines, "fam": cpu_fam}


def generate_kconfig(kconfig, context, overwrite, verbose):
    """Generic kconfig file generator."""
    loader = FileSystemLoader(searchpath=CURRENT_DIR)
    env = Environment(
        loader=loader, trim_blocks=False, lstrip_blocks=True,
        keep_trailing_newline=True
    )
    template_file = os.path.join("Kconfig.{}.j2".format(kconfig))
    env.globals.update(zip=zip)
    template = env.get_template(template_file)
    context.update({"year": datetime.datetime.now().year})
    render = template.render(**context)

    kconfig_dir = os.path.join(STM32_KCONFIG_DIR, context["fam"])
    kconfig_file = os.path.join(kconfig_dir, "Kconfig.{}".format(kconfig))

    # Create family Kconfig dir if it doesn't exist yet
    if not os.path.exists(kconfig_dir):
        os.makedirs(kconfig_dir)

    if (not os.path.exists(kconfig_file) or
            (os.path.exists(kconfig_file) and
             overwrite is True and
             kconfig == "models")):
        with open(kconfig_file, "w") as f_dest:
            f_dest.write(render)
    if verbose:
        print("{}:".format(os.path.basename(kconfig_file)))
        print("-" * (len(os.path.basename(kconfig_file)) + 1) + "\n")
        print(render)


def main(args):
    """Main function."""
    models = parse_sheet(args.cpu_fam, args.sheets)
    lines = parse_cpu_lines(args.cpu_fam)
    context = get_context(args.cpu_fam, models, lines)
    if args.verbose:
        print("Generated kconfig files:\n")
    for kconfig in ["lines", "models"]:
        generate_kconfig(kconfig, context, args.overwrite, args.verbose)


PARSER = argparse.ArgumentParser()
PARSER.add_argument("cpu_fam",
                    help="STM32 CPU Family")
PARSER.add_argument("--sheets", "-s", nargs='+',
                    help="Excel sheet containing the list of products")
PARSER.add_argument("--overwrite", "-o", action="store_true",
                    help="Overwrite any existing Kconfig file")
PARSER.add_argument("--verbose", "-v", action="store_true",
                    help="Print generated file content")


if __name__ == "__main__":
    main(PARSER.parse_args())
